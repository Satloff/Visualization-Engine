#
# Theo Satloff
# Spring 2016
# Data.py v2
#

import csv
import numpy as np
from openpyxl import load_workbook
import time
import datetime


debug = False

class Data:
	def __init__(self, filename1 = None, filename2 = None):
		self.filename1 = filename1
		self.filename2 = filename2
		self.names = [] # list of names
		self.headers = [] # list containing header names
		self.type = [] # list containing data types
		self.data = [] # list containing all of the data read in
		self.header2raw = {} # dictionary mapping headers to the data
		self.header2matrix = {} # dictionary mapping headers to the numeric data
		self.enum = {} # dicitonary that maps strings to enumerated type
		if filename1 is not None:
			self.readIn()
		# if filename2 is not None:
		# 	self.writeOut()


	#-----------------------------------------------------------#
	#------------------- Read data from file--------------------#
	#-----------------------------------------------------------#
	def readIn(self):
		try:
			iterate = None # variable that will hold list to iterate through
			i = 0
			matrixVal = [] # temp list holding the rows for the matrix


			#---------Check Type of file for reading in---------#
			#---------convert csv format into list---------#
			if self.filename1.split('.')[-1] == 'csv': # if it is a .csv
				print "Opened file: ",self.filename1
				f = open(self.filename1, 'rU')
				iterate = csv.reader(f)

			#---------convert xlsx format into list---------#
			elif self.filename1.split('.')[-1] == 'xlsx': # if it is a . xlsx
				print "Opened file: ",self.filename1
				wb = load_workbook(filename = self.filename1)
				sheet_ranges = wb.active

				list = [] # temp list to hold rows of workbook
				for row in sheet_ranges.iter_rows(): # go through each row in the workbook
					temp = [] # temp list to hold single row of workbook
					for cell in row:
						if type(cell.value) == type(1): # if the value is an integer, append
							temp.append(cell.value)
						else:
							temp.append(str(cell.value).encode('ascii', 'replace')) # replace unicode with ascii if not int
					list.append(temp)

				iterate = list
			else:
				print "\n#########################\nError: file type not supported\n#########################\n"
				return

			#---------iterate through the list---------#
			q = 0
			for row in iterate:
				#print row
				if i == 0: # if it is the header row
					self.headers = row
				elif i == 1: # if is the type row
					self.type = row
				else:
					columnVal = [] # temp list to represent each row of matrix
					headerVal = [] # temp list to hold the values of headers used by the matrix
					k = 0

					self.data.append(row) # store all data in data
					for col in range(len(self.headers)):
						if self.type[col] == "numeric": # check to see if the column contains numbers. id normally do .isdigit()
							columnVal.append(float(row[col])) # append value to a temporary list

							if i == 2: # if is the first row of data
								self.header2matrix[self.headers[col]] = k # create dictionary that matches headers with matrix indecies
								k += 1
						elif self.type[col] == "enum":
							value = row[col]
							if value in self.enum:
								if debug:
									print "value already exists"
							else:

								self.enum[value] = q # map value of enum to an int
								q += 1

					matrixVal.append(columnVal) # append the rows of the matrix to self.data
				i += 1
			self.matrixData = np.matrix(matrixVal) # add all of the numeric data to matrix by row


			#---------add entries to header2raw---------#
			j = 0
			for each in self.headers:
				self.header2raw[self.headers[j]] = j
				j += 1

			#ex = self.getNumColRaw()
			#print ex
			#self.getValue(1, "thing1")

			if debug:
				print "Matrix Data: \n", self.matrixData
				print "Data: ", self.data
				print "Header to Raw: ", self.header2raw
				print "Header to Matrix: ", self.header2matrix

		except OSError as e:
			print "\n#########################\nError: please provide a valid read in filename\n#########################\n"


	#-----------------------------------------------------------#
	#--------------------Write list to CSV----------------------#
	#-----------------------------------------------------------#
	def writeOut(self, fileObject, headers, cType = None):
		#print "fileObject: ", fileObject
		#print headers
		data = self.getDataNum(headers)
		temp = []
		temp.insert(0, headers)
		if cType != None:
			value = [cType]*len(headers)

			temp.insert(1, value)

		for i in range(data.shape[0]):
			temp.append(data[i,:].tolist()[0])

		writer = csv.writer(fileObject)
		writer.writerows(temp)


	#-----------------------------------------------------------#
	#--------------------------Getters--------------------------#
	#-----------------------------------------------------------#

	#---------RAW: Get List of Headers---------#
	def getHeaderRaw(self):
		return self.headers

	#---------RAW: Get List of Types---------#
	def getTypesRaw(self):
		return self.type

	#---------RAW: Get Number of Columns---------#
	def getNumColRaw(self):
		return len(self.data[0])

	#---------RAW: Get Number of Rows---------#
	def getNumRowRaw(self):
		return len(self.data)

	#---------RAW: Get a Row from int value---------#
	def getRowRaw(self, int):
		for i, row in enumerate(self.data):
			if i == int:
				return row
			i += 1

	#---------RAW: Get a cell value from int and header---------#
	def getValueRaw(self, rowNum, header):
		for i, each in enumerate(self.headers):
			if each == header:
				break
			i += 1

		for j, row in enumerate(self.data):
			if j == rowNum:
				return row[i]
			j += 1


	#---------NUMERICAL: Get List of Headers---------#
	def getHeaderNum(self):
		temp = []
		dictList = []

		for key, value in self.header2matrix.iteritems():
			dictList.append(key)

		return dictList

	#---------NUMERICAL: Get List of Headers in order---------#
	def getOrderedHeaderNum(self):
		ordered = sorted(self.header2matrix, key=self.header2matrix.__getitem__) # order the dictionary based on the key values
		return ordered

	#---------NUMERICAL: Get Number of Columns---------#
	def getNumColNum(self):
		if debug:
			print "getNumCol: ", len(self.getHeaderNum())
		return len(self.getHeaderNum())

	#---------NUMERICAL: Get Number of Rows---------#
	def getNumRowNum(self):
		if debug:
			print "getNumRow: ", len(self.data)
		return len(self.data)

	#---------NUMERICAL: Get a Row from an int value---------#
	def getRowNum(self, int):
		testVal = int + 1
		if debug:
			print "getRow: ", self.matrixData[int,:].tolist()[0] # get an entire row from the matrix in list form
		return self.matrixData[int,:].tolist()[0]



	#---------NUMERICAL: Get the value of cell from int and header---------#
	def getValueNum(self, int, header):
		val = self.header2matrix[header]
		if debug:
			print "getRow: ", self.matrixData[int,val].tolist() # get a single entry from rox/col in the matrix
		return self.matrixData[int,val].tolist()

	#---------NUMERICAL: Get a matrix from start-end of selected headers---------#
	def getDataNum(self, headerList, start=0, end = 0):
		if end == 0:
			end = self.getNumRowNum()-1

		testVal = start + end

		matrix = []
		for row in range(self.getNumRowNum()):
			if row >= start and row <= end:
				temp = []
				for each in headerList:
					col = self.header2matrix[each]
					temp.append(self.matrixData[row,col])
				matrix.append(temp)
		return np.matrix(matrix)

	#---------General: add a column to the data if supplied header, type, and dimension---------#
	def addCol(self, header, type, list):

		if header in self.headers:
			print "\n#########################\nError: header name already exists\n#########################\n"
			return
		if len(list) == len(self.data):
			for i in range(len(self.data)):
				self.data[i].append(list[i])
			self.headers.append(header)
			self.type.append(type)
		else:
			print "\n#########################\nError: not enough rows\n#########################\n"

		if type == "numeric": # check to see if the column contains numbers. id normally do .isdigit()
			self.header2matrix[header] = len(self.header2matrix) # create dictionary that matches headers with matrix indecies
			a = np.array([list]).transpose()
			self.matrixData = np.append(self.matrixData, a, axis=1)
		#elif self.type[len(self.data)-1] == "enum":
		#	self.enum[row[col]] = q # map value of enum to an int
		#	columnVal.append(float("{0:.1f}".format(q))) # append a float value of 1 decimal place
		#	q += 1

		#	if i == 2: # if is the first row of data
		#		self.header2matrix[self.headers[col]] = k # create dictionary that matches headers with matrix indecies
		#		k += 1

	def main(self):
		print "completed reading in, etc."
		print "see test.py for results"


if __name__ == "__main__":

	dataApp = Data("datasets/drugdeaths.xlsx", "data2.csv")
	dataApp.main()
